"""Crew Gourmet: Excel（マスターシート）→ 公開用 JSON 変換スクリプト。

ローカル作業専用。生のExcelはgit管理対象外（.gitignore）で、
このスクリプトの出力（src/data/restaurants.json）だけを
コミットしてVercelのビルドに使う。

使い方:
    python3 scripts/convert-restaurants.py          # 変換してJSONを書き出す
    python3 scripts/convert-restaurants.py --check  # 書き出さずに検証のみ
"""

from __future__ import annotations

import json
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
EXCEL_PATH = REPO_ROOT / "data" / "crew-gourmet-master.xlsx"
OUTPUT_PATH = REPO_ROOT / "src" / "data" / "restaurants.json"
SHEET_NAME = "マスター"

REQUIRED_COLUMNS = ["No", "店名", "空港", "エリア", "ジャンル"]

# 「飲み」はExcelの専用列「飲み」（○/◎→true、×→false）を最優先し、
# 空欄・要確認の場合のみジャンルの明示語から派生させる（2026-07-19に専用列を新設）。
DRINK_GENRE_KEYWORDS = ["居酒屋", "バー", "パブ", "酒場", "やきとん", "もつ焼き", "日本酒"]

EMOJI_BREAKFAST = "🍳"
EMOJI_LATE_NIGHT = "🌙"
EMOJI_SOLO = "👤"
EMOJI_WALKABLE = "🚶"
EMOJI_CASHLESS = "💳"
EMOJI_QUICK = "⏱️"


class ConversionError(Exception):
    """変換を停止させるべき構造的エラー（必須列欠落・重複など）。"""


@dataclass
class ExclusionInfo:
    id: str
    name: str
    reasons: list[str]


@dataclass
class ConversionReport:
    total_rows: int = 0
    published: list[dict[str, Any]] = field(default_factory=list)
    excluded: list[ExclusionInfo] = field(default_factory=list)
    hours_hidden_ids: list[str] = field(default_factory=list)
    hours_unofficial_ids: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


def normalize_optional(value: Any) -> str | None:
    """空欄・「要確認」を null 化する。それ以外は前後空白を除去した文字列。"""
    if value is None:
        return None
    text = str(value).strip()
    if text == "" or "要確認" in text:
        return None
    return text


def tri_state(value: Any) -> bool | None:
    """○/◎で始まる→true、×で始まる→false、要確認/空欄/「想定」だけの情報→null。"""
    if value is None:
        return None
    text = str(value).strip()
    if text == "" or "要確認" in text:
        return None
    if "想定" in text:
        return None
    if text.startswith("○") or text.startswith("◎"):
        return True
    if text.startswith("×"):
        return False
    return None


def parse_crew_summary(value: Any) -> list[str]:
    """改行で分割し、先頭の「・」を除去して配列化する。原文引用の引用符は付けない。"""
    if value is None:
        return []
    lines = str(value).split("\n")
    result: list[str] = []
    for raw_line in lines:
        line = raw_line.strip()
        if not line:
            continue
        if line.startswith("・"):
            line = line[1:].strip()
        if line:
            result.append(line)
    return result


def is_drink_spot(genre: str) -> bool:
    return any(keyword in genre for keyword in DRINK_GENRE_KEYWORDS)


def format_date(value: Any) -> str | None:
    if value is None:
        return None
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    return text or None


def build_tags(features: dict[str, bool | None], tag_text: str) -> list[str]:
    tags: list[str] = []
    if features["breakfast"] is True:
        tags.append("breakfast")
    if EMOJI_BREAKFAST in tag_text:
        tags.append("earlyMorning")
    if EMOJI_LATE_NIGHT in tag_text:
        tags.append("lateNight")
    if EMOJI_SOLO in tag_text:
        tags.append("solo")
    if EMOJI_WALKABLE in tag_text:
        tags.append("walkable")
    if EMOJI_CASHLESS in tag_text:
        tags.append("cashless")
    if EMOJI_QUICK in tag_text:
        tags.append("quick")
    if features["takeout"] is True:
        tags.append("takeout")
    if features["drink"] is True:
        tags.append("drink")
    return tags


def classify_publish(trust: str | None, status_raw: str) -> list[str]:
    """非公開にすべき理由のリストを返す。空リストなら公開対象。"""
    reasons: list[str] = []
    trust_text = (trust or "").strip()
    status_text = status_raw or ""

    if trust_text == "△":
        reasons.append("信頼度が△のため非公開")
    if "閉店" in status_text or "掲載停止" in status_text:
        reasons.append("閉店・掲載停止のため非公開")
    if "営業状況要確認" in status_text:
        reasons.append("ステータスが「営業状況要確認」のため非公開")
    return reasons


def read_rows(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    if not path.exists():
        raise ConversionError(
            f"Excelファイルが見つかりません: {path}\n"
            "data/crew-gourmet-master.xlsx をローカルへ配置してください。"
        )
    workbook = openpyxl.load_workbook(path, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise ConversionError(f"シート「{SHEET_NAME}」が見つかりません。")
    sheet = workbook[SHEET_NAME]
    header = [cell.value for cell in sheet[1]]

    missing = [col for col in REQUIRED_COLUMNS if col not in header]
    if missing:
        raise ConversionError(
            "必須列が見つからないため変換を停止しました: " + ", ".join(missing)
        )

    rows: list[dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        rows.append(dict(zip(header, row)))
    return header, rows


def convert(path: Path = EXCEL_PATH) -> ConversionReport:
    _, rows = read_rows(path)

    report = ConversionReport(total_rows=len(rows))

    # --- 必須値チェック（行単位） ---
    missing_value_errors: list[str] = []
    for row in rows:
        row_no = row.get("No")
        for col in REQUIRED_COLUMNS:
            value = row.get(col)
            if value is None or str(value).strip() == "":
                missing_value_errors.append(f"No.{row_no}: 「{col}」が空欄です")
    if missing_value_errors:
        raise ConversionError(
            "必須値が空欄の行があるため変換を停止しました:\n"
            + "\n".join(missing_value_errors)
        )

    # --- id重複チェック ---
    seen_ids: dict[str, Any] = {}
    for row in rows:
        row_id = str(row["No"]).strip()
        if row_id in seen_ids:
            raise ConversionError(f"idが重複しています: No.{row_id}")
        seen_ids[row_id] = row

    # --- 店名＋エリア重複チェック ---
    seen_name_area: dict[tuple[str, str], Any] = {}
    for row in rows:
        key = (str(row["店名"]).strip(), str(row["エリア"]).strip())
        if key in seen_name_area:
            raise ConversionError(f"店名＋エリアが重複しています: {key[0]} / {key[1]}")
        seen_name_area[key] = row

    # --- 出典URL形式チェック（警告のみ、内部利用のため出力はしない） ---
    for row in rows:
        url = row.get("出典URL")
        if url is None:
            continue
        url_text = str(url).strip()
        if url_text and not (url_text.startswith("http://") or url_text.startswith("https://")):
            report.warnings.append(
                f"No.{row['No']} {row['店名']}: 出典URLがhttp(s)で始まっていません（{url_text}）"
            )

    for row in rows:
        row_id = str(row["No"]).strip()
        name = str(row["店名"]).strip()
        tag_text = str(row.get("タグ") or "")
        status_raw = str(row.get("ステータス") or "")
        trust = row.get("信頼度")

        drink_explicit = tri_state(row.get("飲み"))
        features = {
            "breakfast": tri_state(row.get("朝食営業")),
            "lateNight": True if EMOJI_LATE_NIGHT in tag_text else (
                False if str(row.get("深夜営業") or "").strip().startswith("×") else None
            ),
            "solo": True if EMOJI_SOLO in tag_text else None,
            "walkable": True if EMOJI_WALKABLE in tag_text else None,
            "takeout": tri_state(row.get("テイクアウト")),
            "cashless": True if EMOJI_CASHLESS in tag_text else None,
            "quick": True if EMOJI_QUICK in tag_text else None,
            "drink": (
                drink_explicit
                if drink_explicit is not None
                else is_drink_spot(str(row.get("ジャンル") or ""))
            ),
        }

        exclusion_reasons = classify_publish(trust, status_raw)

        if exclusion_reasons:
            report.excluded.append(ExclusionInfo(id=row_id, name=name, reasons=exclusion_reasons))
            continue

        hours_uncertain = "営業時間要確認" in status_raw
        if hours_uncertain:
            hours = None
            status_note = "営業時間は最新情報をご確認ください"
            report.hours_hidden_ids.append(row_id)
        else:
            hours = normalize_optional(row.get("営業時間"))
            status_note = None
            if hours and "公式未確認" in hours:
                # 内部注記（例:「（グルメサイト情報・公式未確認）」）は公開画面に出さず、
                # 代わりに公開用の注意文をstatusNoteへ設定する（2026-07-20 管理人指示）。
                hours = re.sub(r"[（(][^（）()]*公式未確認[^（）()]*[）)]", "", hours).strip() or None
                status_note = "営業時間は変更される場合があります。来店前に最新情報をご確認ください。"
                report.hours_unofficial_ids.append(row_id)

        area = str(row["エリア"]).strip()
        airport = str(row["空港"]).strip()

        restaurant = {
            "id": row_id,
            "name": name,
            "airport": airport,
            "area": area,
            "access": normalize_optional(row.get("最寄駅・アクセス")),
            "genre": str(row["ジャンル"]).strip(),
            "specialty": normalize_optional(row.get("名物料理")),
            "crewSummary": parse_crew_summary(row.get("航空関係者の口コミ要約")),
            "hours": hours,
            "closedDays": normalize_optional(row.get("定休日")),
            "features": features,
            "tags": build_tags(features, tag_text),
            "statusNote": status_note,
            "checkedAt": format_date(row.get("最終確認日")),
            "addedAt": format_date(row.get("追加日")),
            "mapQuery": f"{name} {area} {airport}",
        }
        report.published.append(restaurant)

    return report


def print_report(report: ConversionReport, *, check_only: bool) -> None:
    lines = []
    lines.append("=== Crew Gourmet データ変換 ===")
    lines.append(f"読み込み行数: {report.total_rows}")
    lines.append(f"公開対象: {len(report.published)}件")
    lines.append(f"非公開: {len(report.excluded)}件")
    for item in report.excluded:
        lines.append(f"  - No.{item.id} {item.name}: {'; '.join(item.reasons)}")
    if report.hours_hidden_ids:
        lines.append(
            "営業時間を非表示にした店舗（営業時間要確認）: "
            + ", ".join(f"No.{i}" for i in report.hours_hidden_ids)
        )
    if report.hours_unofficial_ids:
        lines.append(
            "営業時間が公式未確認のため注意文を付けた店舗: "
            + ", ".join(f"No.{i}" for i in report.hours_unofficial_ids)
        )
    if report.warnings:
        lines.append("警告:")
        for warning in report.warnings:
            lines.append(f"  - {warning}")

    tag_counts: dict[str, int] = {}
    for restaurant in report.published:
        for tag in restaurant["tags"]:
            tag_counts[tag] = tag_counts.get(tag, 0) + 1
    lines.append("タグ集計（公開対象のみ）:")
    for tag, count in sorted(tag_counts.items()):
        lines.append(f"  - {tag}: {count}件")

    if check_only:
        lines.append("(--check モードのためJSONは書き出していません)")
    else:
        lines.append(f"書き出し先: {OUTPUT_PATH.relative_to(REPO_ROOT)}")

    print("\n".join(lines))


def main() -> int:
    check_only = "--check" in sys.argv[1:]
    try:
        report = convert()
    except ConversionError as error:
        print(f"[エラー] {error}", file=sys.stderr)
        return 1

    print_report(report, check_only=check_only)

    if not check_only:
        OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
        with OUTPUT_PATH.open("w", encoding="utf-8") as f:
            json.dump(report.published, f, ensure_ascii=False, indent=2)
            f.write("\n")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
